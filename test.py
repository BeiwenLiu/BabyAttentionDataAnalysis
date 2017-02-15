#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
Created on Thu Nov 24 21:18:06 2016

@author: Beiwen Liu
"""
import pandas as pd
import numpy as np
import numpy.ma as ma
from scipy.interpolate import interp1d
from scipy.io import loadmat
import matplotlib.pyplot as plt
import numpy
from scipy import interpolate
import pandas as pd
from openpyxl import Workbook
import openpyxl
import os.path
from django.utils.encoding import smart_str
from openpyxl.styles import colors
from openpyxl.styles import Font, Color

data = []

c1 = []
c2 = []
c3 = []
c4 = []
c5 = []
c6 = []
c7 = []
c8 = []
c9 = []

def matCalc():
    
    visitedNumber = -1
    previousNumber = -1
    directory = "mat/"
    whichOne = raw_input("Would you like to analyze just one mat file (1) or all mat files? (2)?\n")
    threshold = raw_input("State your offset from average (usually .2)\n")
    durationthreshold = raw_input("State your duration threshold (0 for now)\n")
    lookdatapoints = raw_input("Minimum number of data points threshold (1 or 2 for now)\n")
    lookdatapoints = int(lookdatapoints)
    if durationthreshold == "":
        durationthreshold = 0
    else:
        durationthreshold = float(durationthreshold)
        
    if (whichOne) == "2":
        for filename in os.listdir(directory):
            previousNumber = visitedNumber
            visitedNumber = filename[-6]
            if filename.endswith(".mat") and (visitedNumber != previousNumber) and (visitedNumber == "2" or visitedNumber == "3" or visitedNumber == "4"):
                main(directory,filename,threshold,durationthreshold,whichOne, lookdatapoints)
                
        createCSV()
    elif (whichOne) == "1":
        tempFile = raw_input("Please type your desired mat file name\n")
        main(directory,tempFile,threshold,durationthreshold,whichOne, lookdatapoints)
    
def createCSV():
    df = pd.DataFrame()
    df['Participants'] = c1
    df['Total Time'] = c9
    df['Median'] = c2 #Median Distance
    df['Mean'] = c3 #Mean Distance
    df['Calculated Distraction Percent'] = c5 #Of all start-stop, how much is baby distracted?
    df['Look Threshold'] = c6 #The threshold that determines a look
    df['Duration Threshold'] = c7 #The threhold that determines a look
    df['Occurrences'] = c8 #Number of start-stop occurrences of look aways
    df.to_csv("allParticipants.csv")
    
def main(directory,filename, threshold,durationthreshold,signal, lookdatapoints):
    if (signal == "1"):
        selectGraph = raw_input("Please select an option:\n(1) Time Series Only\n(2) Graph and Time Series\n(3) Histogram of Distances\n(4) Histogram of Look Durations\n")
    
    mainFile = directory+filename
    
    x = loadmat(mainFile)
    xValues = []
    yValues = []
    timeValues = []

    #print x['EyeData'][4][0]
    for y in range(0,len(x['EyeData'])):
        xValues.append(x['EyeData'][y][4])
        yValues.append(x['EyeData'][y][5])
        timeValues.append(x['EyeData'][y][0])
        
    originalTimeValues = np.ma.compressed(timeValues).tolist()[1:]
    tempInit = originalTimeValues[0]
    originalTimeValues[:] = [z - tempInit for z in originalTimeValues]
    #graphXY(xValues,yValues)
       
    
    interpolateOrNo = raw_input("1) Drop -1\n2) Interpolate\n")
    #Drop all -1 values within X Values 
    if interpolateOrNo == "1":
        
        pre = np.array(xValues)
        sa = np.ma.masked_where(pre == -1, pre)
        
        nXValues = np.ma.masked_array(xValues, sa.mask)
        nTimeValues = np.ma.masked_array(timeValues, sa.mask)
        
        print nTimeValues,"nTime"
        
        xValues = np.ma.compressed(nXValues).tolist()[1:]
        timeValues = np.ma.compressed(nTimeValues).tolist()[1:]
    elif interpolateOrNo == "2":
        pre = inter(xValues)
        
        xValues = np.ma.compressed(pre).tolist()[1:]
        timeValues = np.ma.compressed(timeValues).tolist()[1:]
    
    initialValue = timeValues[0]
    print initialValue, "s"

    timeValues[:] = [x1 - initialValue for x1 in timeValues]
    
    averageX = sum(xValues)/len(xValues)
    medianX = median(xValues)
    
    
        
    #.066 conservative
    #.2 liberal
    
    totalTime = timeValues[-1] - timeValues[0]
    distractedTime = calculateTotalDistractionTime(medianX,threshold,xValues,timeValues)
    
    print "Average X", averageX
    print "Median", medianX
    print totalTime
    print "distraction: " + str(round(distractedTime/totalTime*100,2)) + "%"
    
   
    newTimeValues = []
    newxValues = []
    for i in range(0,len(timeValues)):
        if timeValues[i] not in newTimeValues:
            newTimeValues.append(timeValues[i])
            newxValues.append(xValues[i])
            
    
        
    xStart,yStart,xEnd,yEnd,averageDistance,durations,counter1 = calculations(medianX,threshold,newxValues,newTimeValues,durationthreshold,originalTimeValues, lookdatapoints) #Finds start-end for +x
    xStart1,yStart1,xEnd1,yEnd1,averageDistance1,durations2,counter2 = calculations2(medianX,threshold,newxValues,newTimeValues,durationthreshold, lookdatapoints) #Finds start-end for -x
    
    
    if (selectGraph == "1"):
        windowSize = raw_input("Please Indicate the Window Size\n\n")
        findMovingAverage(int(windowSize),xValues,timeValues, medianX, xStart, xStart1)
        #movingAverage(timeValues,xValues) #Moving average graph
        
    if (selectGraph == "2"):
        graphX(xStart,yStart,xEnd,yEnd,xStart1,yStart1,xEnd1,yEnd1,timeValues,xValues, medianX) #plots everything
    
    durations.extend(durations2) #combine both durations from top and bottom
    if (selectGraph == "3"):
        histogram(xValues) # To enable histogram for distance look away, uncomment here
    if (selectGraph == "4"):
        histogramDurations(durations) # To enable histogram for duration of look aways, uncomment here
    

    total = findDuration(xStart,xEnd) #the total duration of look aways
    print round(total/totalTime*100,2)
    print "distraction: " + str(round(distractedTime/totalTime*100,2)) + "%"
    totalOccurrences = counter1+counter2
    
    
    #The following code will generate statistics
    #saveToExcel(median(xValues),averageX, xStart,xEnd,averageDistance,round(total/totalTime*100,2),round(distractedTime/totalTime*100,2),threshold)
    if (signal == "2"):
        addToLists(filename,totalTime,median(xValues),averageX,round(total/totalTime*100,2),str(round(distractedTime/totalTime*100,2)) + "%",threshold,durationthreshold,totalOccurrences)
    
#Export to CSV of all participants    
def addToLists(filename,totalTime,median,mean,dataDistractionPercent,calculatedDistractionPercent,lookthreshold,durationthreshold,occurrences):
    c1.append(filename)
    c2.append(median)
    c3.append(mean)
    c4.append(dataDistractionPercent)
    c5.append(calculatedDistractionPercent)
    c6.append(lookthreshold)
    c7.append(durationthreshold)
    c8.append(occurrences)
    c9.append(totalTime)

def saveToExcel(median,mean,startTime,endTime,averageXDistance,dataDistractionPercent,calculatedDistractionPercent,threshold):
    new_file = 'results.xlsx'
    wb = openpyxl.Workbook()
    wb.create_sheet(index=0, title='Results') #Creating sheets based on original excel
    ws = wb.worksheets[0]
    ws['A1'] = 'Start Time'
    ws['B1'] = 'End time'
    ws['C1'] = 'Average X Distance'
    ws['D1'] = 'Median Center'
    ws['E1'] = 'Mean Center'
    ws['F1'] = 'Data Distraction Percentage' #Based on found look aways
    ws['G1'] = 'Caulcated Distraction Percentage' #Based on raw data
    ws['H1'] = 'Offset Threshold'
    ws['D2'] = str(median)
    ws['E2'] = str(mean)
    for x in range(0,len(startTime)):
        ws['A{}'.format(x+2)] = str(startTime[x])
    for x in range(0,len(endTime)):
        ws['B{}'.format(x+2)] = str(endTime[x])
    for x in range(0,len(averageXDistance)):
        ws['C{}'.format(x+2)] = str(averageXDistance[x])
    ws['F2'] = str(dataDistractionPercent)
    ws['G2'] = str(calculatedDistractionPercent)
    ws['H2'] = str(threshold)
    wb.save(new_file)

#This method will find the start and end points
def calculations(averageX,threshold,xValues,timeValues,durationthreshold,timeValues1, lookdatapoints):
    threshold = float(threshold)
    
    found = False
    
    currentY = 0
    previousY = xValues[1]
    currentX = 0
    previousX = timeValues[1]
    
    startY = 0 
    startX = 0
    endY = 0
    endX = 0
    
    startXY = []
    endXY = []
    averageDistance = []

    tempDistance = []

    counter = 0
    durations = []
    
    npTime1 = np.array(timeValues)
    npTime2 = np.array(timeValues1) #Original time values with -1
    
    for x in range(2,len(xValues)):
        currentY = xValues[x] # X Distance Value
        currentX = timeValues[x]
        if ~found and (previousY < threshold + averageX and currentY > threshold + averageX):
            startX = slopeX(currentY,previousY,currentX,previousX,threshold + averageX)
            startY = averageX + threshold
            found = True
            
        if found:
            tempDistance.append(currentY)
        
        if found and (previousY > threshold + averageX and currentY < threshold + averageX):
            
            endX = slopeX(currentY,previousY,currentX,previousX,threshold + averageX)
            endY = averageX + threshold
            found = False
            
            constraint = (npTime1 > startX) * (npTime1 < endX)
            constraint2 = (npTime2 > startX) * (npTime2 < endX)
            resultnpTime1 = np.where(constraint)
            resultnpTime2 = np.where(constraint2)
            
            
            '''
            Checks if there is a -1
            goThrough = False
            if npTime1[resultnpTime1].size == npTime2[resultnpTime2].size:
                goThrough = True
            '''
            
            if endX-startX > durationthreshold and npTime1[resultnpTime1].size >= lookdatapoints:
                startXY.append([startX,startY]) #column 0 is the time and column 1 is the "x" distance
                endXY.append([endX,endY])
                durations.append(endX-startX)
                counter = counter + 1
                averageDistance.append(sum(tempDistance) / len(tempDistance))
                
            tempDistance = []
            
        previousY = currentY
        previousX = currentX

    return [row[0] for row in startXY], [row[1] for row in startXY], [row[0] for row in endXY], [row[1] for row in endXY], averageDistance, durations,counter
        
def calculations2(averageX,threshold,xValues,timeValues,durationthreshold, lookdatapoints):
    threshold = float(threshold)
    
    found = False
    
    currentY = 0
    previousY = xValues[1]
    currentX = 0
    previousX = timeValues[1]
    
    startY = 0 
    startX = 0
    endY = 0
    endX = 0
    
    startXY = []
    endXY = []
    averageDistance = []

    tempDistance = []
    
    durations = []
    counter = 0
    
    npTime1 = np.array(timeValues)
    
    for x in range(2,len(xValues)):
        currentY = xValues[x] # X Distance Value
        currentX = timeValues[x]
        if ~found and (previousY > averageX - threshold and currentY < averageX - threshold):
            startX = slopeX(currentY,previousY,currentX,previousX,averageX - threshold)
            startY = averageX - threshold
            found = True
            
        if found:
            tempDistance.append(currentY)
        
        if found and (previousY < averageX - threshold and currentY > averageX - threshold):
            
            endX = slopeX(currentY,previousY,currentX,previousX,averageX - threshold)
            endY = averageX - threshold
            found = False
            
            
            constraint = (npTime1 > startX) * (npTime1 < endX)
            resultnpTime1 = np.where(constraint)
            
            
            
            
            if endX-startX > durationthreshold and npTime1[resultnpTime1].size >= lookdatapoints:
                startXY.append([startX,startY]) #column 0 is the time and column 1 is the "x" distance
                endXY.append([endX,endY])
                durations.append(endX-startX)
                averageDistance.append(sum(tempDistance) / len(tempDistance))
                counter = counter + 1
            tempDistance = []
            
        previousY = currentY
        previousX = currentX

    return [row[0] for row in startXY], [row[1] for row in startXY], [row[0] for row in endXY], [row[1] for row in endXY], averageDistance, durations,counter

def calculateTotalDistractionTime(averageX,threshold,xValues,timeValues):
    distractedTime = 0
    print threshold
    for x in range(1,len(xValues)):
        if (xValues[x] > averageX + float(threshold)) or (xValues[x] < averageX - float(threshold)):
            distractedTime = distractedTime + (timeValues[x]-timeValues[x-1])
            
    return distractedTime
    
def findDuration(start,end):
    duration = 0
    for x in range(0,len(start)):
        duration = duration + (end[0] - start[0])
    return duration
     
#This method will find the associated time with the "x" offset threshold as selected value between two known points       
def slopeX(y2,y1,x2,x1,threshold):
    #print y2,y1,x2,x1,threshold
    slope = (y2 - y1)/(x2 - x1)
    b = y1-slope*x1
    xVal = (threshold - b) / slope
    
    #print "original: " + str(x2)
    #print "new : " + str(xVal)
    return xVal
    
def graphXY(x,y):
    axes = plt.gca()
    plt.scatter(x, y, s=1)
    axes.set_xlim([0,1])
    axes.set_ylim([-2,3])
    plt.show()

def graphX(startX,startY,endX,endY,startX1,startY1,endX1,endY1,time,x,middleX):
    whichOne = raw_input("(1) Add Time Series\n(2) Add Looks\n(3) Add both\n(4) None\n")
    
    middleXList = []
    for num in range(0,len(time)):
        middleXList.append(middleX)
        
    
    if (whichOne == "1" or whichOne == "3"):  
        #following df and graph will plot moving average
        '''
        df = pd.DataFrame(index=time,columns=['Distance Away'])
        df['Distance Away'] = x
        graph = df.rolling(window=1000,center=False).mean()
        graph.plot(style='bs-')'''
        windowSize = raw_input("Please Indicate the Window Size\n\n")
        findMovingAverage(int(windowSize),x,time,middleX, startX, startX1,endX,endX1)
        
    print len(time)
    print len(x)
    df = pd.DataFrame(columns=['a'])
    df['a']=time
    df['b']=x
    df.to_csv('test.csv')
    
    plt.plot(time,x)
    plt.plot(time,middleXList, 'r') #Graphs center line according to middleX which can be average or median
    
    if (whichOne == "2" or whichOne == "3"):
        plt.scatter(startX,startY)
        plt.scatter(startX1,startY1)
        plt.scatter(endX,endY)
        plt.scatter(endX1,endY1)
    #axes.set_xlim([0,1])
    #axes.set_ylim([-2,3])
    plt.show()
    
def ase():
    plt.axis([0, 10, 0, 1])
    plt.ion()

    for i in range(10):
        y = np.random.random()
        plt.scatter(i, y)
        plt.pause(0.05)

    while True:
        plt.pause(0.05)
        

def median(lst):
    return numpy.median(numpy.array(lst))

def practice():
    list1 = []
    list1.append([1,2])
    list1.append([2,3])
    print list1[0][0]

#Histogram for distribution of distances 
def histogram(sa):
    
    # Get histogram of random data
    bins = []
    x = .1
    while x < 1:
        bins.append(x)
        x += .01
        
    y, x = np.histogram(sa, bins=bins)
    
    # Correct bin placement
    x = x[1:]
    
    # Turn into pandas Series
    hist = pd.Series(y, x)
    
    # Plot
    ax = hist.plot(kind='bar', width=1, alpha=0.5, align='center')
    ax.set_title('Uniform Bin Distribution of Distance Look Away')
    ax.set_xlabel('Distance Look Away')
    
#Histogram for distribution of durations 
def histogramDurations(sa):
    
    sa = [int(round(x)) for x in sa]
    minimum = min(sa)
    maximum = max(sa)
    binSize = (maximum-minimum) / 100
    print "max",maximum,"min",minimum, "bin",binSize
    
    # Get histogram of random data
    bins = []
    x = minimum
    while x <= maximum:
        bins.append(x)
        x += binSize
        
    y, x = np.histogram(sa, bins=bins)
    
    # Correct bin placement
    x = x[1:]
    
    # Turn into pandas Series
    hist = pd.Series(y, x)
    
    # Plot
    ax = hist.plot(kind='bar', width=1, alpha=0.5, align='center')
    ax.set_title('Uniform Bin Distribution of Look Durations')
    ax.set_xlabel('Duration of Look Away')

def movingAverage(timeValues,xValues):
    df = pd.DataFrame(index=timeValues,columns=['Distance Away'])
    df['Distance Away'] = xValues
    graph = df.rolling(window=1000,center=False).mean()
    graph.plot(style='r')
    
def inter(xValues):
    pre = np.array(xValues)
    
    pre[pre==-1] = numpy.nan

    nans, x= nan_helper(pre)
    pre[nans]= np.interp(x(nans), x(~nans), pre[~nans])
    
    return pre

def nan_helper(y):
    """Helper to handle indices and logical indices of NaNs.

    Input:
        - y, 1d numpy array with possible NaNs
    Output:
        - nans, logical indices of NaNs
        - index, a function, with signature indices= index(logical_indices),
          to convert logical indices of NaNs to 'equivalent' indices
    Example:
        >>> # linear interpolation of NaNs
        >>> nans, x= nan_helper(y)
        >>> y[nans]= np.interp(x(nans), x(~nans), y[~nans])
    """

    return np.isnan(y), lambda z: z.nonzero()[0]
    
def findMovingAverage(windowSize,xValues,timeValues,median,xStart, xStart1, xEnd, xEnd1):
    newXVals = []
    distanceOrInstance = raw_input("1) Distance\n2) Instance\n3) Look Distance")
    if distanceOrInstance == "1":
        xValues = [abs(median - k) + median for k in xValues] #Median minus distance, absolute value, add median for offset
        for x in range(0,len(xValues) - windowSize):
            newXVals.append(sum(xValues[x:x+windowSize])/windowSize) 
        df = pd.DataFrame(index=timeValues[:-windowSize],columns=['Distance Away'])
        df['Distance Away'] = newXVals
        df.plot(style='r')
    elif distanceOrInstance == "2":
        print xStart[-1], xEnd[-1]
        xValues = [abs(median - k) + median for k in xValues] #Median minus distance, absolute value, add median for offset
        newXVals= []
        
        for x in range(windowSize / 2,len(xValues) - (windowSize / 2)):
            #newXVals.append(sum(xValues[x - windowSize/2 : x+windowSize/2])/windowSize)
            counter = 0
            counter2 = 0
            total = 0
            numberOccurrence = 0
            for y in range(x - windowSize/2,x + windowSize/2):
                if counter < len(xEnd) and timeValues[y] > xStart[counter] and timeValues[y] < xEnd[counter]:
                    total = total + xValues[y]
                    numberOccurrence = numberOccurrence + 1
                else:
                    while counter < len(xEnd) and timeValues[y] > xEnd[counter]:
                        counter = counter + 1
                if counter2 < len(xEnd1) and timeValues[y] > xStart1[counter2] and timeValues[y] < xEnd1[counter2]:
                    total = total + xValues[y]
                    numberOccurrence = numberOccurrence + 1
                else:
                    while counter2 < len(xEnd1) and timeValues[y] > xEnd1[counter2]:
                        counter2 = counter2 + 1
            if numberOccurrence != 0:
                newXVals.append(total / numberOccurrence)
                print total/numberOccurrence
            else:
                newXVals.append(median)
                print median


        df = pd.DataFrame(index=timeValues[windowSize/2:-windowSize/2],columns=['Distance Away'])
        df['Distance Away'] = newXVals
        df.plot(style='r')
        
    
        
#ase()
#main()
#data_listener()
#practice()

matCalc()
#inter()